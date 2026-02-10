"""
Ingest routes: bulk load issues from Excel (RAVE/QC export style).

POST /ingest/issues accepts a single .xlsx file and creates issues via the same
normalizer used by the CLI script.
"""

from __future__ import annotations

import io
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile

from agent.ingest.normalizers import from_excel_row
from apps.api import storage

router = APIRouter(prefix="/ingest", tags=["ingest"])

MAX_ROWS = 200
MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5 MB


def _read_excel_rows(content: bytes) -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2):
        cells = [c.value for c in row]
        row_dict: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if i < len(cells) and h:
                v = cells[i]
                row_dict[h] = v
        if any(v is not None and v != "" for v in row_dict.values()):
            rows.append(row_dict)
    wb.close()
    return rows


@router.post("/issues")
def ingest_issues_from_excel(file: UploadFile = File(...)) -> dict[str, Any]:
    """
    Upload an Excel file (.xlsx) to bulk-create issues.

    Expected columns: Source, Domain, Subject_ID, Fields, Description;
    optional: Start_Date, End_Date, Variable, Value, Reference, Notes.
    Rows that fail validation are skipped and reported in `errors`.
    """
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx files are accepted",
        )

    content = file.file.read()
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large (max {MAX_UPLOAD_BYTES // (1024*1024)} MB)",
        )

    rows = _read_excel_rows(content)
    if len(rows) > MAX_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"Too many data rows (max {MAX_ROWS})",
        )

    issue_ids: list[str] = []
    errors: list[str] = []
    for i, row in enumerate(rows):
        try:
            issue_create = from_excel_row(row)
            issue = storage.BACKEND.create_issue(issue_create)
            issue_ids.append(str(issue.issue_id))
        except Exception as e:
            errors.append(f"Row {i + 2}: {e!s}")

    return {
        "created": len(issue_ids),
        "issue_ids": issue_ids,
        "errors": errors,
    }
