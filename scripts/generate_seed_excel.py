"""
Generate data/seed/rave_export_demo.xlsx for demo ingestion.

Run from repo root: python scripts/generate_seed_excel.py

Creates 10 rows of commonly seen drug-development issues (RAVE/QC style).
Column schema matches agent/ingest/normalizers.from_excel_row and the plan.
"""

from __future__ import annotations

import sys
from pathlib import Path

# Repo root
ROOT = Path(__file__).resolve().parent.parent
SEED_DIR = ROOT / "data" / "seed"
OUT_PATH = SEED_DIR / "rave_export_demo.xlsx"

HEADERS = [
    "Source",
    "Domain",
    "Subject_ID",
    "Fields",
    "Description",
    "Start_Date",
    "End_Date",
    "Variable",
    "Value",
    "Reference",
    "Notes",
]

ROWS = [
    # These issues will be automatically classified by the system
    # Deterministic examples (simple, rule-based)
    [
        "edit_check",
        "AE",
        "SUBJ-1001",
        "AESTDTC, AEENDTC",
        "AE end date is before start date.",
        "2024-01-15",
        "2024-01-10",
        "",
        "",
        "AE_DATE_001",
        "RAVE edit check",
    ],
    [
        "edit_check",
        "AE",
        "SUBJ-1002",
        "AESER, AEOUT",
        "Missing required field AESER for serious AE.",
        "",
        "",
        "AESER",
        "",
        "AESER required when AESDTH=Y",
        "",
    ],
    [
        "listing",
        "LB",
        "SUBJ-1003",
        "LBORRES, LBSTRESN",
        "Lab value out of range: hemoglobin.",
        "",
        "",
        "LBORRES",
        "4.2",
        "Normal 12-17 g/dL",
        "Visit V2",
    ],
    [
        "listing",
        "VS",
        "SUBJ-1004",
        "VSORRES, VSSUPYN",
        "Vital sign out of range: systolic BP 180 mmHg.",
        "",
        "",
        "VSORRES",
        "180",
        "Normal 90-140 mmHg",
        "",
    ],
    [
        "listing",
        "AE",
        "SUBJ-1005",
        "AESTDTC, AETERM",
        "Duplicate record suspected: same subject, date, event.",
        "2024-02-01",
        "2024-02-01",
        "AETERM",
        "Headache",
        "Duplicate key",
        "",
    ],
    [
        "edit_check",
        "DM",
        "SUBJ-1006",
        "RFSTDTC, VISIT",
        "Visit date outside protocol-defined window.",
        "2024-03-15",
        "",
        "VISIT",
        "V3",
        "Window 14-21 days from V2",
        "",
    ],
    [
        "edit_check",
        "AE",
        "SUBJ-1008",
        "AEENDTC",
        "Missing end date for ongoing AE.",
        "2024-01-20",
        "",
        "AEENDTC",
        "",
        "Required when AEOUT=Ongoing",
        "",
    ],
    # LLM-required examples (complex, requires nuanced analysis)
    [
        "listing",
        "LB",
        "SUBJ-1009",
        "LBORRES, LBCLSIG",
        "Discrepancy vs central lab: EDC value differs from external. Clinical significance unclear.",
        "",
        "",
        "LBORRES",
        "12.5",
        "Central lab 11.8",
        "LBCLSIG=Y, requires clinical context",
    ],
    [
        "edit_check",
        "DM",
        "SUBJ-1010",
        "BRTHDTC",
        "Invalid or partial date: birth date month/year only. May be acceptable per protocol.",
        "",
        "",
        "BRTHDTC",
        "1985-03",
        "ISO 8601 full date",
        "Partial date may be acceptable",
    ],
    [
        "listing",
        "AE",
        "SUBJ-1011",
        "AETERM, AESEV, AESER",
        "Complex adverse event with multiple related conditions. Requires medical review to determine if single or multiple events.",
        "2024-04-10",
        "",
        "AETERM",
        "Headache, Nausea, Dizziness",
        "Multiple symptoms",
        "Need to assess if related or separate",
    ],
    [
        "listing",
        "DM",
        "SUBJ-1012",
        "DMWEIGHT, DMWTU",
        "Inconsistent units: weight in kg vs lb across visits. Need to assess impact on BMI calculations.",
        "",
        "",
        "DMWTU",
        "lb",
        "Study standard kg",
        "Previous visit was kg, BMI affected",
    ],
    [
        "edit_check",
        "AE",
        "SUBJ-1013",
        "AESTDTC, AEENDTC, AESER",
        "Serious adverse event with ambiguous timeline. Start date conflicts with hospitalization records.",
        "2024-05-01",
        "2024-05-15",
        "AESTDTC",
        "2024-05-03",
        "Hospitalization started 2024-05-03",
        "Date reconciliation needed",
    ],
    # Issue without RAG guidance (for testing "no guidance found" scenario)
    [
        "listing",
        "CM",
        "SUBJ-1014",
        "CMTRT, CMDOSFRM",
        "Uncommon medication coding issue: combination product not in standard dictionary. Requires manual review.",
        "",
        "",
        "CMTRT",
        "Product-XY Plus",
        "Not in MedDRA/WHODrug",
        "Novel combination product",
    ],
]


def main() -> int:
    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        return 1

    SEED_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        return 1
    ws.title = "Issues"

    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=c, value=h)

    for r, row in enumerate(ROWS, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH} ({len(ROWS)} rows)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
