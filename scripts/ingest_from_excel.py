"""
Ingest issues from an Excel file (RAVE/QC export style) by POSTing to the API.

Usage (from repo root):
  python scripts/ingest_from_excel.py [path_to.xlsx] [--base-url URL] [--dry-run]

Default path: data/seed/rave_export_demo.xlsx
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
DEFAULT_EXCEL = ROOT / "data" / "seed" / "rave_export_demo.xlsx"
DEFAULT_BASE_URL = "http://127.0.0.1:8000"


def _load_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


def _read_sheet_rows(path: Path) -> list[dict[str, str | int | float | None]]:
    openpyxl = _load_openpyxl()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows: list[dict[str, str | int | float | None]] = []
    for row in ws.iter_rows(min_row=2):
        cells = [c.value for c in row]
        row_dict = {}
        for i, h in enumerate(headers):
            if i < len(cells) and h:
                v = cells[i]
                if isinstance(v, (str, int, float)):
                    row_dict[h] = v
                elif v is not None:
                    row_dict[h] = str(v)
                else:
                    row_dict[h] = None
        if any(v is not None and v != "" for v in row_dict.values()):
            rows.append(row_dict)
    wb.close()
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description="Ingest issues from Excel to API")
    parser.add_argument(
        "path",
        nargs="?",
        type=Path,
        default=DEFAULT_EXCEL,
        help=f"Path to .xlsx file (default: {DEFAULT_EXCEL})",
    )
    parser.add_argument(
        "--base-url",
        default=DEFAULT_BASE_URL,
        help=f"API base URL (default: {DEFAULT_BASE_URL})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate and print would-be payloads without POSTing",
    )
    args = parser.parse_args()

    path = args.path
    if not path.is_file():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    try:
        from agent.ingest.normalizers import from_excel_row
    except ImportError:
        print("Run from repo root so agent.ingest is importable", file=sys.stderr)
        return 1

    rows = _read_sheet_rows(path)
    if not rows:
        print("No data rows in sheet")
        return 0

    creates: list[dict] = []
    errors: list[str] = []
    for i, row in enumerate(rows):
        try:
            issue_create = from_excel_row(row)
            creates.append(issue_create.model_dump(mode="json"))
        except Exception as e:
            errors.append(f"Row {i + 2}: {e}")

    if errors:
        for e in errors:
            print(e, file=sys.stderr)
        if not creates:
            return 1

    if args.dry_run:
        print(f"Dry run: would POST {len(creates)} issues to {args.base_url}/issues")
        for j, payload in enumerate(creates[:3]):
            print(f"  [{j+1}] {payload.get('description', '')[:50]}...")
        if len(creates) > 3:
            print(f"  ... and {len(creates) - 3} more")
        return 0

    try:
        import httpx
    except ImportError:
        print("Install httpx: pip install httpx", file=sys.stderr)
        return 1

    issue_ids: list[str] = []
    for payload in creates:
        try:
            r = httpx.post(
                f"{args.base_url.rstrip('/')}/issues",
                json=payload,
                timeout=30.0,
            )
            r.raise_for_status()
            data = r.json()
            issue_ids.append(str(data.get("issue_id", "")))
            print(data.get("issue_id"))
        except Exception as e:
            errors.append(f"POST failed for {payload.get('description', '')[:30]}: {e}")

    if errors and not issue_ids:
        for e in errors:
            print(e, file=sys.stderr)
        return 1
    print(f"Created {len(issue_ids)} issues.", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
